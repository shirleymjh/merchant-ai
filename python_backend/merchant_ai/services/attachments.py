from __future__ import annotations

import json
import shutil
import time
import uuid
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Any, Iterable

from merchant_ai.config import Settings
from merchant_ai.services.text_parsing import is_ascii_hex


class AttachmentStore:
    def __init__(self, settings: Settings):
        self.settings = settings
        self.root = settings.resolved_workspace_path / "attachments"
        self.root.mkdir(parents=True, exist_ok=True)
        self.cleanup_expired(int(settings.attachment_retention_days or 7))

    def save(self, name: str, content_type: str, payload: bytes, merchant_id: str = "") -> dict[str, Any]:
        attachment_id = "attachment_" + uuid.uuid4().hex
        safe_name = safe_filename(name)
        directory = self.root / attachment_id
        directory.mkdir(parents=True, exist_ok=True)
        target = directory / safe_name
        target.write_bytes(payload)
        preview, parser = extract_text_preview(
            safe_name,
            content_type,
            payload,
            max_chars=int(self.settings.attachment_preview_max_chars or 12000),
        )
        metadata = {
            "attachmentId": attachment_id,
            "name": safe_name,
            "type": content_type or "application/octet-stream",
            "size": len(payload),
            "path": str(target),
            "textPreview": preview,
            "parser": parser,
            "merchantId": str(merchant_id or self.settings.merchant_id),
            "createdAt": int(time.time()),
        }
        (directory / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
        return metadata

    def context_for(self, references: Iterable[Any], merchant_id: str = "") -> str:
        sections: list[str] = []
        for reference in references or []:
            attachment_id = str(getattr(reference, "id", "") or "")
            if not _is_attachment_id(attachment_id):
                continue
            metadata_path = self.root / attachment_id / "metadata.json"
            try:
                metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            owner = str(metadata.get("merchantId") or "")
            if merchant_id and owner and owner != str(merchant_id):
                continue
            preview = str(metadata.get("textPreview") or "")[: int(self.settings.attachment_preview_max_chars or 12000)]
            summary = "附件：%s（%s，%s bytes）" % (
                metadata.get("name") or attachment_id,
                metadata.get("type") or "unknown",
                metadata.get("size") or 0,
            )
            parser = str(metadata.get("parser") or "none")
            sections.append(
                summary
                + (("\n解析方式：%s\n可解析内容：\n%s" % (parser, preview)) if preview else "\n附件未提取到可读文本，请提示用户换用清晰图片或可复制文本的文档。")
            )
        return "\n\n".join(sections)

    def cleanup_expired(self, retention_days: int) -> None:
        cutoff = time.time() - max(1, retention_days) * 86400
        for directory in self.root.glob("attachment_*"):
            try:
                if directory.is_dir() and directory.stat().st_mtime < cutoff:
                    shutil.rmtree(directory)
            except OSError:
                continue


def safe_filename(name: str) -> str:
    value = Path(str(name or "attachment")).name
    value = "".join(
        character if character.isalnum() or character in {"_", ".", "-"} else "_"
        for character in value
    )
    return value[:160] or "attachment"


def _is_attachment_id(value: str) -> bool:
    prefix = "attachment_"
    suffix = value[len(prefix) :] if value.startswith(prefix) else ""
    return len(suffix) == 32 and suffix == suffix.lower() and is_ascii_hex(suffix, minimum=32)


_OCR_ENGINE: Any = None
_OCR_LOCK = Lock()


def extract_text_preview(name: str, content_type: str, payload: bytes, max_chars: int = 12000) -> tuple[str, str]:
    suffix = Path(name).suffix.lower()
    if str(content_type or "").startswith("text/") or name.lower().endswith((".csv", ".txt", ".json", ".md")):
        for encoding in ("utf-8-sig", "utf-8", "gb18030"):
            try:
                return payload.decode(encoding)[:max_chars], "text"
            except UnicodeDecodeError:
                continue
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return extract_excel_preview(payload, max_chars), "excel"
    if suffix == ".xls":
        return extract_legacy_excel_preview(payload, max_chars), "excel"
    if suffix == ".pdf" or content_type == "application/pdf":
        return extract_pdf_preview(payload, max_chars), "pdf"
    if str(content_type or "").startswith("image/") or suffix in {".png", ".jpg", ".jpeg", ".webp", ".bmp"}:
        return extract_image_ocr(payload, max_chars), "ocr"
    return "", "unsupported"


def extract_excel_preview(payload: bytes, max_chars: int) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    output: list[str] = []
    try:
        for worksheet in workbook.worksheets[:8]:
            output.append("工作表：%s" % worksheet.title)
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_number > 200:
                    output.append("（该工作表仅展示前 200 行）")
                    break
                values = [normalize_cell(value) for value in row[:40]]
                while values and not values[-1]:
                    values.pop()
                if values:
                    output.append("\t".join(values))
                if sum(len(line) + 1 for line in output) >= max_chars:
                    return "\n".join(output)[:max_chars]
    finally:
        workbook.close()
    return "\n".join(output)[:max_chars]


def extract_legacy_excel_preview(payload: bytes, max_chars: int) -> str:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=payload, on_demand=True)
    output: list[str] = []
    try:
        for sheet_name in workbook.sheet_names()[:8]:
            worksheet = workbook.sheet_by_name(sheet_name)
            output.append("工作表：%s" % sheet_name)
            for row_number in range(min(worksheet.nrows, 200)):
                values = [
                    normalize_cell(worksheet.cell_value(row_number, column))
                    for column in range(min(worksheet.ncols, 40))
                ]
                while values and not values[-1]:
                    values.pop()
                if values:
                    output.append("\t".join(values))
                if sum(len(line) + 1 for line in output) >= max_chars:
                    return "\n".join(output)[:max_chars]
    finally:
        workbook.release_resources()
    return "\n".join(output)[:max_chars]


def extract_pdf_preview(payload: bytes, max_chars: int) -> str:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(payload))
    pages: list[str] = []
    for index, page in enumerate(reader.pages[:30], start=1):
        text = str(page.extract_text() or "").strip()
        if text:
            pages.append("第 %s 页\n%s" % (index, text))
        if sum(len(item) + 1 for item in pages) >= max_chars:
            break
    return "\n\n".join(pages)[:max_chars]


def extract_image_ocr(payload: bytes, max_chars: int) -> str:
    global _OCR_ENGINE
    from PIL import Image
    from rapidocr_onnxruntime import RapidOCR

    with _OCR_LOCK:
        if _OCR_ENGINE is None:
            _OCR_ENGINE = RapidOCR()
        image = Image.open(BytesIO(payload)).convert("RGB")
        if max(image.size) > 2400:
            image.thumbnail((2400, 2400))
        result, _ = _OCR_ENGINE(image)
    lines = [str(item[1]).strip() for item in (result or []) if len(item) > 1 and str(item[1]).strip()]
    return "\n".join(lines)[:max_chars]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return text[:500]
